
import re
import io
import json
import zipfile
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import streamlit as st

st.set_page_config(page_title="Multi-Category Heatmap (Compact)", layout="wide")

# ------------------------ CONFIG: CATEGORIES & QUESTIONS ------------------------
CATEGORIES = {
    "Spend Visibility & Control": [
        ("Q1: Consolidated real-time view of total spend (Qualitative)", ""),
        ("Q2: Spend Under Management % (KPI, Best ≥80%)", ""),
        ("Q3: Maverick Spend % (KPI, Best ≤5%)", ""),
        ("Q4: Detailed categorization & consistent tagging (Qualitative)", ""),
        ("Q5: Contract Compliance % (KPI, Best ≥90%)", ""),
    ],
    "Vendor & Supply Chain Mgmt": [
        ("Q1: Top-10 supplier concentration & dual/multi-sourcing (KPI/Qual)", ""),
        ("Q2: OTIF – On-Time, In-Full % (KPI)", ""),
        ("Q3: Supplier performance reviews tied to C/Q/D & action plans (Qual)", ""),
        ("Q4: % spend under contracts with cost-reduction clauses/value-add (KPI)", ""),
        ("Q5: Total Landed Cost used in sourcing/renewals (Qual)", ""),
    ],
    "Process & Workforce Efficiency": [
        ("Q1: Touchless/straight-through processing rate (KPI)", ""),
        ("Q2: Cost per transaction (AP invoice, service ticket, etc.) (KPI)", ""),
        ("Q3: First-pass yield / rework rate (KPI)", ""),
        ("Q4: Cycle time for key processes (P2P, O2C, monthly close) (KPI)", ""),
        ("Q5: Workforce utilization & overtime control (Qual/KPI)", ""),
    ],
    "Production & Asset Utilization": [
        ("Q1: OEE – Overall Equipment Effectiveness % (KPI)", ""),
        ("Q2: Unplanned downtime as % of scheduled time (KPI)", ""),
        ("Q3: % assets under preventive/predictive maintenance (KPI)", ""),
        ("Q4: Inventory health (Turns or Days on Hand) (KPI)", ""),
        ("Q5: Capacity utilization vs plan (KPI/Qual)", ""),
    ],
    "Energy & Facility Costs": [
        ("Q1: Energy intensity (kWh per output unit/per sqft) tracked & improved (KPI)", ""),
        ("Q2: % energy spend under hedged/contracted rates (KPI)", ""),
        ("Q3: Facility occupancy/utilization rate (KPI)", ""),
        ("Q4: Data center PUE / IT energy KPI (KPI)", ""),
        ("Q5: Waste/water cost control & recycling/reuse programs (Qual/KPI)", ""),
    ],
}

# Columns common to all categories
COLUMNS = ["Procurement", "Finance", "IT", "Operations"]

# Subtle, professional colors to match your screenshot
# 1=Red, 3=Yellow, 5=Green (not too saturated)
COLOR_MAP = {
    1: "#FF9999",  # soft red
    3: "#FFF59D",  # soft yellow
    5: "#A5D6A7",  # soft green
}
SCORES = [1, 3, 5]  # simple dropdown

# Compact sizing to avoid extra space
Q_COL_W = "520px"     # Questions column width (same as your screenshot)
CELL_W  = "90px"      # Score cell width
CELL_H  = "36px"      # Row height
FONT    = "system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial, sans-serif"

# File outputs
EXCEL_PATH = "multi_category_heatmaps.xlsx"
CSV_PREFIX = "heatmap_"

# ------------------------ HTML renderer (compact, clean) ------------------------
def render_html_table(df: pd.DataFrame, title: str) -> str:
    rows = df.values.tolist()
    index = df.index.tolist()
    columns = df.columns.tolist()

    css = f'''
    <style>
      .hm-wrap {{ font-family:{FONT}; color:#1f2937; }}
      .hm-title {{ font-weight:700; margin:0 0 6px 0; }}
      .hm-table {{
        border-collapse: collapse;
        border:1px solid #d1d5db; /* gray-300 */
        width: fit-content;
        font-size:13px;
      }}
      .hm-table th, .hm-table td {{
        border:1px solid #e5e7eb; /* gray-200 */
        width:{CELL_W}; height:{CELL_H};
        text-align:center; vertical-align:middle; padding:0;
      }}
      .hm-table th:first-child, .hm-table td:first-child {{
        width:{Q_COL_W}; text-align:left; padding:6px 10px;
        white-space:normal; line-height:1.3;
      }}
      .hm-table th {{ background:#f7f7f7; font-weight:600; }}
    </style>
    '''
    html = [f'<div class="hm-wrap">']
    html.append(f'<div class="hm-title">{title}</div>')
    html.append('<table class="hm-table">')

    # Header
    html.append('<tr>')
    html.append('<th>Questions</th>')
    for c in columns:
        html.append(f'<th>{c}</th>')
    html.append('</tr>')

    # Rows
    for i, row in enumerate(rows):
        q = index[i]
        html.append('<tr>')
        html.append(f'<td title="{q}">{q}</td>')
        for val in row:
            bg = COLOR_MAP.get(int(val), "#ffffff")
            html.append(f'<td style="background:{bg};">{int(val)}</td>')
        html.append('</tr>')

    html.append('</table></div>')
    return css + "\n".join(html)

def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[\[\]\*\?\\/]', '_', name)
    return name[:31]

def to_excel_bytes(category_to_df: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for cat, df in category_to_df.items():
        ws = wb.create_sheet(title=sanitize_sheet_name(cat))
        export_df = df.reset_index().rename(columns={"index": "Questions"})
        for r in dataframe_to_rows(export_df, index=False, header=True):
            ws.append(r)
        # Apply fills to data cells only
        for row in ws.iter_rows(min_row=2, min_col=2, max_row=1+len(df.index), max_col=1+len(df.columns)):
            for cell in row:
                v = cell.value
                if isinstance(v, (int, float)) and int(v) in COLOR_MAP:
                    hex6 = COLOR_MAP[int(v)].lstrip("#")
                    cell.fill = PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def to_csv_zip_bytes(category_to_df: dict) -> bytes:
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for cat, df in category_to_df.items():
            fname = f"{CSV_PREFIX}{cat.replace(' ', '_').replace('&', 'and')}.csv"
            zf.writestr(fname, df.to_csv(index_label="Questions"))
    mem.seek(0)
    return mem.getvalue()

# ------------------------ State helpers ------------------------
def init_state():
    if "scores" not in st.session_state:
        st.session_state.scores = {}
        for cat, items in CATEGORIES.items():
            st.session_state.scores[cat] = {}
            for q, _ in items:
                st.session_state.scores[cat][q] = {c: 3 for c in COLUMNS}
    if "last_action" not in st.session_state:
        st.session_state["last_action"] = "one"
    if "last_cat" not in st.session_state:
        st.session_state["last_cat"] = list(CATEGORIES.keys())[0]

def get_df_for_category(cat: str) -> pd.DataFrame:
    rows = []
    questions = [q for q, _ in CATEGORIES[cat]]
    for q in questions:
        rows.append([st.session_state.scores[cat][q][c] for c in COLUMNS])
    return pd.DataFrame(rows, index=questions, columns=COLUMNS)

# ------------------------ UI ------------------------
init_state()

st.title("Multi-Category Heatmap — Compact, Subtle Colors, Simple 1/3/5")
st.caption("Inputs on the left • Heatmap output on the right • Save to Excel or CSVs")

left, right = st.columns([0.55, 0.45], gap="large")

with left:
    st.subheader("Inputs")
    cat = st.selectbox("Category", list(CATEGORIES.keys()), index=list(CATEGORIES.keys()).index(st.session_state["last_cat"]))
    st.session_state["last_cat"] = cat
    st.markdown("Fill simple **1 / 3 / 5** scores per function.")

    # Build grid of dropdowns
    for (q, _) in CATEGORIES[cat]:
        cols = st.columns([1.8] + [1,1,1,1])
        cols[0].markdown(f"**{q}**")
        for i, c in enumerate(COLUMNS):
            key = f"{cat}|{q}|{c}"
            current = st.session_state.scores[cat][q][c]
            st.session_state.scores[cat][q][c] = cols[i+1].selectbox(
                c, [1,3,5],
                index=[1,3,5].index(current) if current in [1,3,5] else 1,
                key=key
            )

    st.markdown("---")
    gen = st.button("Generate Category Heatmap", type="primary", use_container_width=True)
    gen_all = st.button("Generate All Categories", use_container_width=True)

with right:
    st.subheader("Output")
    if gen:
        st.session_state["last_action"] = "one"
    if gen_all:
        st.session_state["last_action"] = "all"

    if st.session_state["last_action"] == "one":
        df = get_df_for_category(st.session_state["last_cat"])
        html = render_html_table(df, title=st.session_state["last_cat"])
        st.components.v1.html(html, height=min(900, 140 + 42*len(df.index)), scrolling=True)
    else:
        for c in CATEGORIES.keys():
            df = get_df_for_category(c)
            html = render_html_table(df, title=c)
            st.components.v1.html(html, height=min(900, 140 + 42*len(df.index)), scrolling=True)

    st.markdown("---")
    # Build data for saving
    category_to_df = {c: get_df_for_category(c) for c in CATEGORIES.keys()}
    excel_bytes = to_excel_bytes(category_to_df)
    csvzip_bytes = to_csv_zip_bytes(category_to_df)
    st.download_button("Download Excel (multi-sheet)", data=excel_bytes, file_name="multi_category_heatmaps.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    st.download_button("Download CSVs (zip)", data=csvzip_bytes, file_name="heatmaps_csvs.zip", mime="application/zip", use_container_width=True)

st.caption("Tip: Deploy on Streamlit Community Cloud or Hugging Face Spaces. The layout keeps inputs and outputs clearly separated for client demos.")
