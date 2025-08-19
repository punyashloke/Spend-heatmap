
import streamlit as st
import pandas as pd
import io, re, zipfile
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------- THEME (subtle, executive) ----------------------------
FONT_STACK = "Inter, system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial, sans-serif"
BORDER_OUTER = "#d1d5db"  # gray-300
BORDER_INNER = "#e5e7eb"  # gray-200
HEADER_BG = "#f7f7f7"
TEXT_COLOR = "#1f2937"    # gray-800

# Softer tones (AA accessible on white)
COLOR_MAP = {  # 1=Red, 3=Yellow, 5=Green
    1: "#FCA5A5",  # soft red-300
    3: "#FDE68A",  # soft amber-300
    5: "#86EFAC",  # soft green-300
}

# Layout constants
Q_COL_W = "520px"  # left column for long questions
CELL_W  = "90px"
CELL_H  = "36px"

st.set_page_config(page_title="Executive Heatmap (Compact)", layout="wide")

# --------------------------------- DATA MODEL -------------------------------------
CATEGORIES = {
    "Spend Visibility & Control": [
        ("Q1: Consolidated real-time view of total spend (Qualitative)", ""),
        ("Q2: Spend Under Management % (KPI, Best ≥80%)", ""),
        ("Q3: Maverick Spend % (KPI, Best ≤5%)", ""),
        ("Q4: Detailed categorization & consistent tagging (Qualitative)", ""),
        ("Q5: Contract Compliance % (KPI, Best ≥90%)", ""),
    ],
    "Vendor & Supply Chain Mgmt": [
        ("Q1: Top-10 supplier concentration & dual/multi-sourcing (KPI/Qual)", ""),
        ("Q2: OTIF – On-Time, In-Full % (KPI)", ""),
        ("Q3: Supplier performance reviews tied to C/Q/D & action plans (Qual)", ""),
        ("Q4: % spend under contracts with cost-reduction clauses/value-add (KPI)", ""),
        ("Q5: Total Landed Cost used in sourcing/renewals (Qual)", ""),
    ],
    "Process & Workforce Efficiency": [
        ("Q1: Touchless/straight-through processing rate (KPI)", ""),
        ("Q2: Cost per transaction (AP invoice, service ticket, etc.) (KPI)", ""),
        ("Q3: First-pass yield / rework rate (KPI)", ""),
        ("Q4: Cycle time for key processes (P2P, O2C, monthly close) (KPI)", ""),
        ("Q5: Workforce utilization & overtime control (Qual/KPI)", ""),
    ],
    "Production & Asset Utilization": [
        ("Q1: OEE – Overall Equipment Effectiveness % (KPI)", ""),
        ("Q2: Unplanned downtime as % of scheduled time (KPI)", ""),
        ("Q3: % assets under preventive/predictive maintenance (KPI)", ""),
        ("Q4: Inventory health (Turns or Days on Hand) (KPI)", ""),
        ("Q5: Capacity utilization vs plan (KPI/Qual)", ""),
    ],
    "Energy & Facility Costs": [
        ("Q1: Energy intensity (kWh per output unit/per sqft) tracked & improved (KPI)", ""),
        ("Q2: % energy spend under hedged/contracted rates (KPI)", ""),
        ("Q3: Facility occupancy/utilization rate (KPI)", ""),
        ("Q4: Data center PUE / IT energy KPI (KPI)", ""),
        ("Q5: Waste/water cost control & recycling/reuse programs (Qual/KPI)", ""),
    ],
    # You can add more categories in the same format if needed
}

COLUMNS = ["Procurement", "Finance", "IT", "Operations"]
SCORES = [1, 3, 5]  # simple dropdowns

# --------------------------------- HELPERS ----------------------------------------
def render_html_table(df: pd.DataFrame, title: str) -> str:
    rows = df.values.tolist()
    index = df.index.tolist()
    columns = df.columns.tolist()
    css = f'''
    <style>
      .hm-wrap {{ font-family:{FONT_STACK}; color:{TEXT_COLOR}; }}
      .hm-title {{ font-weight:700; margin:0 0 6px 0; }}
      .hm-table {{ border-collapse: collapse; border:1px solid {BORDER_OUTER}; width: fit-content; font-size:13px; }}
      .hm-table th, .hm-table td {{ border:1px solid {BORDER_INNER}; width:{CELL_W}; height:{CELL_H}; text-align:center; vertical-align:middle; padding:0; }}
      .hm-table th:first-child, .hm-table td:first-child {{ width:{Q_COL_W}; text-align:left; padding:6px 10px; white-space:normal; line-height:1.3; }}
      .hm-table th {{ background:{HEADER_BG}; font-weight:600; }}
    </style>'''
    # Header
    header = '<tr><th>Questions</th>' + ''.join(f'<th>{c}</th>' for c in columns) + '</tr>'
    # Body
    body_rows = []
    for i, row in enumerate(rows):
        q = index[i]
        tds = ''.join(f'<td style="background:{COLOR_MAP.get(int(val), "#ffffff")};">{int(val)}</td>' for val in row)
        body_rows.append(f'<tr><td title="{q}">{q}</td>{tds}</tr>')
    table = '<table class="hm-table">' + header + ''.join(body_rows) + '</table>'
    return css + f'<div class="hm-wrap"><div class="hm-title">{title}</div>{table}</div>'

def sanitize_sheet_name(name: str) -> str:
    return re.sub(r'[\[\]\*\?\\/]', '_', name)[:31]

@st.cache_data
def excel_bytes(category_to_df: dict) -> bytes:
    bio = io.BytesIO()
    wb = Workbook(); wb.remove(wb.active)
    for cat, df in category_to_df.items():
        ws = wb.create_sheet(title=sanitize_sheet_name(cat))
        export_df = df.reset_index().rename(columns={"index": "Questions"})
        for r in dataframe_to_rows(export_df, index=False, header=True): ws.append(r)
        for row in ws.iter_rows(min_row=2, min_col=2, max_row=1+len(df.index), max_col=1+len(df.columns)):
            for cell in row:
                v = cell.value
                if isinstance(v, (int, float)) and int(v) in COLOR_MAP:
                    hex6 = COLOR_MAP[int(v)].lstrip("#")
                    cell.fill = PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")
    wb.save(bio); bio.seek(0); return bio.getvalue()

@st.cache_data
def csv_zip_bytes(category_to_df: dict) -> bytes:
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for cat, df in category_to_df.items():
            fname = f"heatmap_{cat.replace(' ', '_').replace('&', 'and')}.csv"
            zf.writestr(fname, df.to_csv(index_label="Questions"))
    mem.seek(0); return mem.getvalue()

def init_state():
    if "scores" not in st.session_state:
        st.session_state.scores = {cat: {q: {c:3 for c in COLUMNS} for (q,_) in items} for cat, items in CATEGORIES.items()}
    st.session_state.setdefault("last_action", "one")
    st.session_state.setdefault("last_cat", list(CATEGORIES.keys())[0])

def df_for(cat: str) -> pd.DataFrame:
    qs = [q for (q,_) in CATEGORIES[cat]]
    rows = [[st.session_state.scores[cat][q][c] for c in COLUMNS] for q in qs]
    return pd.DataFrame(rows, index=qs, columns=COLUMNS)

# ----------------------------------- UI ------------------------------------------
init_state()
st.title("Executive Heatmap — Compact, Subtle, Board-Ready")
st.caption("Inputs on the left • Heatmap output on the right • 1 / 3 / 5 only • Excel & CSV downloads")

left, right = st.columns([0.55, 0.45], gap="large")
with left:
    st.subheader("Inputs")
    cat = st.selectbox("Category", list(CATEGORIES.keys()), index=list(CATEGORIES.keys()).index(st.session_state["last_cat"]))
    st.session_state["last_cat"] = cat
    st.write("Set scores for each function: **1 / 3 / 5**")
    for (q,_) in CATEGORIES[cat]:
        cols = st.columns([1.8,1,1,1,1])
        cols[0].markdown(f"**{q}**")
        for i,c in enumerate(COLUMNS):
            st.session_state.scores[cat][q][c] = cols[i+1].selectbox(c, [1,3,5], index=[1,3,5].index(st.session_state.scores[cat][q][c]), key=f"{cat}|{q}|{c}")
    st.markdown("---")
    gen = st.button("Generate Category Heatmap", type="primary", use_container_width=True)
    gen_all = st.button("Generate All Categories", use_container_width=True)

with right:
    st.subheader("Output")
    if gen: st.session_state["last_action"]="one"
    if gen_all: st.session_state["last_action"]="all"
    if st.session_state["last_action"]=="one":
        df = df_for(st.session_state["last_cat"])
        html = render_html_table(df, st.session_state["last_cat"])
        st.components.v1.html(html, height=min(900, 140 + 42*len(df.index)), scrolling=True)
    else:
        for c in CATEGORIES.keys():
            df = df_for(c)
            html = render_html_table(df, c)
            st.components.v1.html(html, height=min(900, 140 + 42*len(df.index)), scrolling=True)

    st.markdown("---")
    cat2df = {c: df_for(c) for c in CATEGORIES.keys()}
    st.download_button("Download Excel (multi-sheet)", data=excel_bytes(cat2df), file_name="multi_category_heatmaps.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    st.download_button("Download CSVs (zip)", data=csv_zip_bytes(cat2df), file_name="heatmaps_csvs.zip", mime="application/zip", use_container_width=True)
